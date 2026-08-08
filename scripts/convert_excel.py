import openpyxl, json, datetime

wb = openpyxl.load_workbook('E:/download/hhhh.xlsx', data_only=True)

def dt_str(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m-%d %H:%M')
    if isinstance(v, datetime.date):
        return v.strftime('%Y-%m-%d')
    return str(v)

# ========== 1. 每日寄样 -> creators + samples ==========
ws = wb['每日寄样']
creators_map = {}
samples = []

for row in ws.iter_rows(min_row=2, values_only=True):
    name = str(row[0]).strip() if row[0] else None
    if not name:
        continue

    videos = []
    for i in range(18, min(45, len(row)), 2):
        vlink = row[i] if i < len(row) else None
        vtime = row[i + 1] if i + 1 < len(row) else None
        if vlink:
            # 飞书导出的超链接可能是 {link, text} 对象或字符串
            if isinstance(vlink, dict):
                vlink = vlink.get('link') or vlink.get('url') or vlink.get('text') or ''
            vlink = str(vlink).strip()
            if vlink:
                videos.append({'url': vlink, 'time': dt_str(vtime)})

    sample = {
        'creator': name,
        'official': str(row[1]) if row[1] else None,
        'stars': str(row[2]) if row[2] else None,
        'creatorType': str(row[3]) if row[3] else None,
        'orderCount': str(row[4]) if row[4] else None,
        'reinvest': str(row[5]) if row[5] else None,
        'approval': str(row[6]) if row[6] else None,
        'language': str(row[7]) if row[7] else None,
        'bodyType': str(row[8]) if row[8] else None,
        'age': str(row[9]) if row[9] else None,
        'fulfillment': str(row[10]) if row[10] else None,
        'category': str(row[11]) if row[11] else None,
        'note': str(row[12]) if row[12] else None,
        'color': str(row[13]) if row[13] else None,
        'sku': str(row[14]) if row[14] else None,
        'sampleTime': dt_str(row[15]),
        'updateTime': dt_str(row[16]),
        'fulfillMethod': str(row[17]) if row[17] else None,
        'videos': videos,
    }
    samples.append(sample)

    if name not in creators_map:
        creators_map[name] = {
            'name': name,
            'official': sample['official'],
            'stars': sample['stars'],
            'creatorType': sample['creatorType'],
            'bodyType': sample['bodyType'],
            'age': sample['age'],
            'fulfillment': sample['fulfillment'],
            'category': sample['category'],
            'note': sample['note'],
            'sampleCount': 0,
            'videoCount': 0,
            'skus': [],
            'lastSampleTime': None,
        }
    c = creators_map[name]
    c['sampleCount'] += 1
    c['videoCount'] += len(videos)
    if sample['sku'] and sample['sku'] not in c['skus']:
        c['skus'].append(sample['sku'])
    if sample['official'] and not c['official']:
        c['official'] = sample['official']
    if sample['stars'] and not c['stars']:
        c['stars'] = sample['stars']
    if sample['sampleTime']:
        if not c['lastSampleTime'] or sample['sampleTime'] > c['lastSampleTime']:
            c['lastSampleTime'] = sample['sampleTime']

creators = list(creators_map.values())

# ========== 2. 暂定寄样达人 ==========
ws2 = wb['暂定寄样达人']
pending = []
for row in ws2.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    pending.append({
        'name': str(row[0]),
        'official': str(row[1]) if row[1] else None,
        'stars': str(row[2]) if row[2] else None,
        'bodyType': str(row[3]) if row[3] else None,
        'age': str(row[4]) if row[4] else None,
        'fulfillment': str(row[5]) if row[5] else None,
        'category': str(row[6]) if row[6] else None,
        'note': str(row[7]) if row[7] else None,
        'color': str(row[8]) if row[8] else None,
        'sku': str(row[9]) if row[9] else None,
        'rejectReason': str(row[10]) if row[10] else None,
        'reprocessTime': dt_str(row[11]),
        'result': str(row[12]) if row[12] else None,
    })

# ========== 3. 邀约达人 ==========
ws3 = wb['邀约达人']
invites = []
for row in ws3.iter_rows(min_row=2, values_only=True):
    if not row[0] and not row[1]:
        continue
    invites.append({
        'date': dt_str(row[0]),
        'sku': str(row[1]) if row[1] else None,
        'creatorId': str(row[2]) if row[2] else None,
        'category': str(row[3]) if row[3] else None,
        'commission': str(row[4]) if row[4] else None,
        'validDays': str(row[5]) if row[5] else None,
        'planTarget': row[6],
        'achieved': row[7],
        'note': str(row[8]) if row[8] else None,
        'script': str(row[9]) if row[9] else None,
    })

# ========== 4. 本月寄样任务 ==========
ws4 = wb['本月寄样任务完成情况']
tasks = []
for row in ws4.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    tasks.append({
        'sku': str(row[0]),
        'priority': str(row[1]) if row[1] else None,
        'positioning': str(row[2]) if row[2] else None,
        'startTime': dt_str(row[3]),
        'target': row[4],
        'completed': row[5],
        'uncompleted': row[6],
        'avgDaily': row[7],
        'todayCompleted': row[8],
        'note': str(row[9]) if row[9] else None,
    })

# ========== 5. 产品定位 ==========
ws5 = wb['产品定位']
skus = []
for row in ws5.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    skus.append({
        'sku': str(row[0]),
        'positioning': str(row[1]) if row[1] else None,
        'productId': str(row[2]) if row[2] else None,
    })

# ========== Output ==========
data = {
    'creators': creators,
    'samples': samples,
    'pending': pending,
    'invites': invites,
    'tasks': tasks,
    'skus': skus,
}

stats = {
    'creatorCount': len(creators),
    'sampleCount': len(samples),
    'pendingCount': len(pending),
    'inviteCount': len(invites),
    'taskCount': len(tasks),
    'skuCount': len(skus),
    'videoCount': sum(len(s['videos']) for s in samples),
}

js_content = '// Auto-generated from Feishu export (hhhh.xlsx)\n'
js_content += '// Generated: 2026-08-02\n'
js_content += '// DO NOT EDIT MANUALLY - regenerate from Excel\n\n'
js_content += 'window.REAL_DATA = ' + json.dumps(data, ensure_ascii=False, indent=2) + ';\n'
js_content += '\nwindow.REAL_DATA_STATS = ' + json.dumps(stats, ensure_ascii=False, indent=2) + ';\n'

with open('C:/Users/LENOVO/WorkBuddy/2026-08-02-13-13-32/js/real-data.js', 'w', encoding='utf-8') as f:
    f.write(js_content)

print('Done! Stats:')
print(json.dumps(stats, indent=2, ensure_ascii=False))
